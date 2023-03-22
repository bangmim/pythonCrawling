import openpyxl

# 1) 엑셀 만들기
wb = openpyxl.Workbook()

# 2) 엑셀 워크시트 만들기 (시트추가)
ws = wb.create_sheet('오징어게임')

# 3) 데이터 추가하기
ws['A1'] = "참가번호"
ws['B1'] = "성명"

ws['A2'] = 1
ws['B2'] = "오일남"

# 4) 엑셀 저장하기 (r + 경로 복사 >> 지정 경로에 저장이된다)
# wb.save('참가자_data.xlsx')       //루트폴더에 저장된다
wb.save(r'D:\새 폴더\인프런\STARTCODING CRAWLING\02_파이썬 엑셀 다루기\참가자_data.xlsx')