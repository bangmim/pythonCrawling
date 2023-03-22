import openpyxl

fpath=r'D:\새 폴더\인프런\STARTCODING CRAWLING\02_파이썬 엑셀 다루기\주식현재가_data.xlsx'

'''
wb= openpyxl.Workbook()
ws = wb.create_sheet('주식현재가')

ws['A1']="종목"
ws['B1']="현재가"
ws['C1']="평균매입가"
ws['D1']="잔고수량"
ws['E1']="평가금액"
ws['F1']="평가손익"
ws['G1']="수익률"

ws['A2']="삼성전자"
ws['B2']=""
ws['C2']=85000
ws['D2']=20
ws['E2']=""
ws['F2']=""
ws['G2']="-100%"

ws['A3']="SK하이닉스"
ws['B3']=""
ws['C3']=120000
ws['D3']=15
ws['E3']=""
ws['F3']=""
ws['G3']="-100%"

ws['A4']="카카오"
ws['B4']=""
ws['C4']=145000
ws['D4']=10
ws['E4']=""
ws['F4']=""
ws['G4']="-100%"

wb.save(fpath)
'''
wb = openpyxl.load_workbook(fpath)
ws = wb['주식현재가']
